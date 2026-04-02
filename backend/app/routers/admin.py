import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Header
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook
from email_validator import validate_email, EmailNotValidError

from app.database import get_db
from app.models import Employee
from app.schemas import AdminLogin, EmployeeCreate, EmployeeOut, BulkUploadResult
from app.config import ADMIN_PASSWORD

router = APIRouter(prefix="/api/admin", tags=["admin"])


def verify_admin(x_admin_password: str = Header(..., alias="X-Admin-Password")):
    if x_admin_password != ADMIN_PASSWORD:
        raise HTTPException(status_code=403, detail="Contraseña incorrecta")


@router.post("/login")
async def admin_login(data: AdminLogin):
    """Verify admin password."""
    if data.password != ADMIN_PASSWORD:
        raise HTTPException(status_code=403, detail="Contraseña incorrecta")
    return {"message": "Acceso concedido"}


@router.get("/employees", response_model=list[EmployeeOut])
async def list_all_employees(
    _=Depends(verify_admin),
    db: AsyncSession = Depends(get_db)
):
    """List all employees (including inactive)."""
    result = await db.execute(
        select(Employee).order_by(Employee.hierarchy_level, Employee.name)
    )
    return result.scalars().all()


@router.post("/employees", response_model=EmployeeOut)
async def create_employee(
    data: EmployeeCreate,
    _=Depends(verify_admin),
    db: AsyncSession = Depends(get_db)
):
    """Create a new employee."""
    employee = Employee(**data.model_dump())
    db.add(employee)
    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(status_code=400, detail=f"Ya existe un empleado con email {data.email}")
    await db.refresh(employee)
    return employee


@router.put("/employees/{employee_id}", response_model=EmployeeOut)
async def update_employee(
    employee_id: int,
    data: EmployeeCreate,
    _=Depends(verify_admin),
    db: AsyncSession = Depends(get_db)
):
    """Update an employee."""
    employee = await db.get(Employee, employee_id)
    if not employee:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    for key, value in data.model_dump().items():
        setattr(employee, key, value)
    await db.commit()
    await db.refresh(employee)
    return employee


@router.patch("/employees/{employee_id}/toggle")
async def toggle_employee(
    employee_id: int,
    _=Depends(verify_admin),
    db: AsyncSession = Depends(get_db)
):
    """Toggle active/inactive status of an employee."""
    employee = await db.get(Employee, employee_id)
    if not employee:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    employee.active = not employee.active
    await db.commit()
    return {"message": f"Empleado {'activado' if employee.active else 'desactivado'}"}


@router.delete("/employees/{employee_id}")
async def delete_employee(
    employee_id: int,
    _=Depends(verify_admin),
    db: AsyncSession = Depends(get_db)
):
    """Permanently delete an employee."""
    employee = await db.get(Employee, employee_id)
    if not employee:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    await db.delete(employee)
    await db.commit()
    return {"message": "Empleado eliminado permanentemente"}


@router.post("/employees/bulk", response_model=BulkUploadResult)
async def bulk_upload(
    file: UploadFile = File(...),
    _=Depends(verify_admin),
    db: AsyncSession = Depends(get_db)
):
    """
    Bulk upload employees from Excel (.xlsx).
    Expected columns: nombre, email, cargo, departamento, nivel_jerarquia
    """

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel (.xlsx)")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="El archivo está vacío")

    # Find header row
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    expected = {"nombre", "email", "cargo", "departamento", "nivel_jerarquia"}
    header_map = {}
    for i, h in enumerate(headers):
        if h in expected:
            header_map[h] = i

    missing = expected - set(header_map.keys())
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Columnas faltantes: {', '.join(missing)}. "
                   f"Esperadas: nombre, email, cargo, departamento, nivel_jerarquia"
        )

    # Phase 1: Validate all rows first
    validated = []
    errors = []

    for row_num, row in enumerate(rows[1:], start=2):
        try:
            name = str(row[header_map["nombre"]]).strip()
            email = str(row[header_map["email"]]).strip()
            position = str(row[header_map["cargo"]]).strip()
            department = str(row[header_map["departamento"]]).strip()
            level = int(row[header_map["nivel_jerarquia"]])

            if not name or not email:
                errors.append(f"Fila {row_num}: nombre o email vacío")
                continue

            try:
                validate_email(email, check_deliverability=False)
            except EmailNotValidError:
                errors.append(f"Fila {row_num}: email inválido '{email}'")
                continue

            if level not in (1, 2, 3, 4):
                errors.append(f"Fila {row_num}: nivel_jerarquia debe ser 1-4")
                continue

            validated.append(dict(
                name=name, email=email, position=position,
                department=department, hierarchy_level=level
            ))
        except Exception as e:
            errors.append(f"Fila {row_num}: {str(e)}")

    wb.close()

    if errors:
        return BulkUploadResult(created=0, errors=errors)

    # Phase 2: Insert all validated rows in a single transaction
    for data in validated:
        db.add(Employee(**data))
    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        return BulkUploadResult(created=0, errors=["Email duplicado encontrado. Ningún empleado fue creado."])

    return BulkUploadResult(created=len(validated), errors=[])


@router.get("/template")
async def download_template():
    """Return the expected Excel template format."""
    return {
        "columns": ["nombre", "email", "cargo", "departamento", "nivel_jerarquia"],
        "hierarchy_levels": {
            1: "Gerente General / Director Ejecutivo",
            2: "Gerentes (Administración y Finanzas, Operaciones)",
            3: "Jefaturas (11 áreas)",
            4: "Empleados que reportan a jefaturas",
        },
        "example_row": {
            "nombre": "Juan Pérez",
            "email": "jperez@faymex.cl",
            "cargo": "Jefe de Bodega",
            "departamento": "Bodega",
            "nivel_jerarquia": 3,
        }
    }
